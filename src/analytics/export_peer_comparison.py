"""Sprint 3 Day 20 - peer_comparison.xlsx: 11 sheets, percentile colour-coded, benchmark highlighted."""
import sqlite3
import pandas as pd
from openpyxl.styles import PatternFill
from src.analytics.peer_engine import compute_peer_percentiles, METRICS

DB = "db/nifty100.db"
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")
GOLD = PatternFill("solid", fgColor="FFD966")


def _fill_for(pct):
    if pd.isna(pct):
        return None
    if pct >= 75:
        return GREEN
    if pct >= 25:
        return YELLOW
    return RED


def export():
    perc = compute_peer_percentiles()
    c = sqlite3.connect(DB)
    peer_groups = pd.read_sql("SELECT * FROM peer_groups", c)
    comp = pd.read_sql("SELECT id AS company_id, company_name FROM companies", c)
    c.close()

    with pd.ExcelWriter("output/peer_comparison.xlsx", engine="openpyxl") as xw:
        for group_name, members in peer_groups.groupby("peer_group_name"):
            members = members.merge(comp, on="company_id", how="left")
            wide_val = perc[perc["peer_group_name"] == group_name].pivot(
                index="company_id", columns="metric", values="value").reset_index()
            wide_pct = perc[perc["peer_group_name"] == group_name].pivot(
                index="company_id", columns="metric", values="percentile_rank").reset_index()
            wide_pct.columns = ["company_id"] + [f"{m}_percentile" for m in wide_pct.columns[1:]]
            out = members[["company_id", "company_name", "is_benchmark"]].merge(wide_val, on="company_id", how="left")
            out = out.merge(wide_pct, on="company_id", how="left")

            median_row = {"company_id": "MEDIAN", "company_name": "Peer Group Median", "is_benchmark": 0}
            for m in METRICS:
                if m in out.columns:
                    median_row[m] = out[m].median()
            out = pd.concat([out, pd.DataFrame([median_row])], ignore_index=True)

            sheet = group_name[:31]
            out.to_excel(xw, sheet_name=sheet, index=False)
            ws = xw.sheets[sheet]
            cols = list(out.columns)
            for r in range(2, len(out) + 2):
                is_bench = out.iloc[r - 2]["is_benchmark"] == 1
                if is_bench:
                    for ci in range(1, len(cols) + 1):
                        ws.cell(row=r, column=ci).fill = GOLD
                for m in METRICS:
                    pcol = f"{m}_percentile"
                    if pcol in cols:
                        ci = cols.index(pcol) + 1
                        val = out.iloc[r - 2][pcol]
                        fill = _fill_for(val)
                        if fill and not is_bench:
                            ws.cell(row=r, column=ci).fill = fill
    print("Wrote output/peer_comparison.xlsx")


if __name__ == "__main__":
    export()