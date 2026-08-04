"""Sprint 3 Day 17 - export screener_output.xlsx, 6 preset sheets, colour-coded."""
import yaml
import pandas as pd
from openpyxl.styles import PatternFill
from src.screener.engine import load_universe, composite_quality_score, apply_filters

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")

COLS = ["company_id", "company_name", "broad_sector", "composite_quality_score",
        "return_on_equity_pct", "debt_to_equity", "free_cash_flow_cr", "revenue_cagr_5yr",
        "pat_cagr_5yr", "operating_profit_margin_pct", "pe_ratio", "pb_ratio",
        "dividend_yield_pct", "interest_coverage", "market_cap_crore", "net_profit_cr",
        "eps_cagr_5yr", "asset_turnover", "sales_cr", "dividend_payout_ratio_pct"]

_META_OP = {">=": lambda v, t: v >= t, "<=": lambda v, t: v <= t, "==": lambda v, t: v == t}


def _threshold_col_and_op(fkey, config):
    special = {
        "max_dividend_payout": ("dividend_payout_ratio_pct", "<="),
        "min_rev_cagr_3yr": ("revenue_cagr_3yr", ">="),
        "fcf_positive_latest": ("free_cash_flow_cr", ">="),
        "de_declining_yoy": (None, None),
    }
    if fkey in special:
        return special[fkey]
    meta = config["metrics"][fkey]
    return meta["column"], meta["op"]


def export():
    with open("config/screener_config.yaml") as f:
        cfg = yaml.safe_load(f)
    universe = composite_quality_score(load_universe())

    with pd.ExcelWriter("output/screener_output.xlsx", engine="openpyxl") as xw:
        for pkey, preset in cfg["presets"].items():
            filtered = apply_filters(universe, preset["filters"], cfg)
            present_cols = [c for c in COLS if c in filtered.columns]
            out = filtered[present_cols].round(2)
            sheet = preset["name"][:31]
            out.to_excel(xw, sheet_name=sheet, index=False)
            ws = xw.sheets[sheet]
            for fkey, val in preset["filters"].items():
                if fkey in ("fcf_positive_latest", "de_declining_yoy"):
                    continue
                col, op = _threshold_col_and_op(fkey, cfg)
                if col not in present_cols:
                    continue
                col_idx = present_cols.index(col) + 1
                for row_i, v in enumerate(out[col], start=2):
                    cell = ws.cell(row=row_i, column=col_idx)
                    passed = _META_OP.get(op, lambda a, b: True)(v, val) if pd.notna(v) else False
                    cell.fill = GREEN if passed else RED
    print("Wrote output/screener_output.xlsx")


if __name__ == "__main__":
    export()